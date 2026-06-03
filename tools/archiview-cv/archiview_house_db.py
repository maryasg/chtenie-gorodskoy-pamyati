#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Archiview CV — отдельный модуль вкладки «База домов».

Модуль импортирует CSV/XLSX/JSON со списком домов, создаёт стандартную папку
проекта дома и отдаёт выбранный дом в основной GUI через callback.
"""
from __future__ import annotations

import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tkinter as tk
from dataclasses import asdict, dataclass, field
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence

STATUS_OPTIONS = [
    "Найден акт",
    "Нужны исторические фото",
    "Нужно современное фото",
    "Готово к выпрямлению",
    "Размечается",
    "Проверено",
    "Опубликовано",
]

ALIASES: Dict[str, Sequence[str]] = {
    "district": ["направление", "район", "округ", "district", "area"],
    "address": ["адрес", "адрес / группа адресов", "адрес/группа адресов", "дом", "address", "house"],
    "object_name": ["объект", "объект / что указано в акте", "объект/что указано в акте", "наименование объекта", "название объекта", "object_name", "object", "name"],
    "act_title": ["тип / тема экспертизы", "тип/тема экспертизы", "тема экспертизы", "тип", "название акта", "act_title", "document", "type"],
    "act_date": ["дата", "дата / период", "дата/период", "период", "год", "act_date", "date", "year"],
    "act_url": ["источник", "ссылка", "ссылка на акт", "url", "pdf", "act_url", "source", "link"],
    "priority": ["приоритет", "приоритет для нашей разметки", "priority"],
    "status": ["статус", "status"],
    "notes": ["комментарий", "комментарии", "примечание", "примечания", "notes", "comment"],
    "lat": ["lat", "latitude", "широта"],
    "lon": ["lon", "lng", "longitude", "долгота"],
    "uid": ["id", "uid", "код", "номер", "record_id"],
    "project_slug": ["project_slug", "slug", "папка", "имя папки", "папка проекта"],
}

CYR = str.maketrans({
    "а":"a","б":"b","в":"v","г":"g","д":"d","е":"e","ё":"e","ж":"zh","з":"z","и":"i","й":"y","к":"k","л":"l","м":"m","н":"n","о":"o","п":"p","р":"r","с":"s","т":"t","у":"u","ф":"f","х":"h","ц":"ts","ч":"ch","ш":"sh","щ":"sch","ъ":"","ы":"y","ь":"","э":"e","ю":"yu","я":"ya",
})

@dataclass
class HouseRecord:
    uid: str = ""
    address: str = ""
    district: str = ""
    object_name: str = ""
    act_title: str = ""
    act_date: str = ""
    act_url: str = ""
    priority: str = ""
    status: str = "Найден акт"
    notes: str = ""
    lat: Optional[float] = None
    lon: Optional[float] = None
    project_slug: str = ""
    source_file: str = ""
    extra: Dict[str, Any] = field(default_factory=dict)

    @property
    def display_title(self) -> str:
        return self.address or self.object_name or self.uid or "Дом без названия"

    def ensure_uid(self, index: int = 0) -> None:
        if not self.uid:
            self.uid = safe_slug(self.address or self.object_name or f"house_{index+1}")
        if not self.project_slug:
            self.project_slug = safe_slug(self.address or self.uid)

    def to_json(self) -> Dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_json(cls, data: Dict[str, Any]) -> "HouseRecord":
        known = set(cls.__dataclass_fields__.keys())  # type: ignore[attr-defined]
        kwargs = {k: v for k, v in data.items() if k in known}
        rec = cls(**kwargs)
        rec.lat = parse_float_or_none(rec.lat)
        rec.lon = parse_float_or_none(rec.lon)
        return rec


def norm_header(x: Any) -> str:
    return re.sub(r"\s+", " ", str(x or "").strip().lower().replace("ё", "е"))


def get_alias(row: Dict[str, Any], field_name: str, default: str = "") -> Any:
    aliases = {norm_header(a) for a in ALIASES[field_name]}
    for k, v in row.items():
        if norm_header(k) in aliases:
            return v
    return default


def parse_float_or_none(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def safe_slug(text: str, max_len: int = 90) -> str:
    s = str(text or "").strip().lower().translate(CYR)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return (s or "house_project")[:max_len]


def read_house_table(path: str | Path) -> List[HouseRecord]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Файл не найден: {p}")
    if p.suffix.lower() in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows(p)
    elif p.suffix.lower() in {".csv", ".txt"}:
        rows = _read_csv_rows(p)
    elif p.suffix.lower() == ".json":
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            data = data.get("records", [])
        records = [HouseRecord.from_json(x) for x in data]
        for i, r in enumerate(records):
            r.source_file = str(p)
            r.ensure_uid(i)
        return records
    else:
        raise ValueError("Поддерживаются .xlsx, .csv, .json")
    return rows_to_records(rows, str(p))


def _read_csv_rows(path: Path) -> List[Dict[str, Any]]:
    last_error: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t") if sample.strip() else csv.excel
                return list(csv.DictReader(f, dialect=dialect))
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"Не удалось прочитать CSV: {last_error}")


def _read_xlsx_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("Для Excel нужен openpyxl. Добавьте в requirements_archiview.txt: openpyxl>=3.1") from exc
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        headers = [str(x or "").strip() for x in next(it)]
    except StopIteration:
        return []
    rows: List[Dict[str, Any]] = []
    for values in it:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
        if any(str(v or "").strip() for v in row.values()):
            rows.append(row)
    return rows


def rows_to_records(rows: Iterable[Dict[str, Any]], source_file: str = "") -> List[HouseRecord]:
    recs: List[HouseRecord] = []
    for i, row in enumerate(rows):
        rec = HouseRecord(
            uid=str(get_alias(row, "uid", "") or "").strip(),
            address=str(get_alias(row, "address", "") or "").strip(),
            district=str(get_alias(row, "district", "") or "").strip(),
            object_name=str(get_alias(row, "object_name", "") or "").strip(),
            act_title=str(get_alias(row, "act_title", "") or "").strip(),
            act_date=str(get_alias(row, "act_date", "") or "").strip(),
            act_url=str(get_alias(row, "act_url", "") or "").strip(),
            priority=str(get_alias(row, "priority", "") or "").strip(),
            status=str(get_alias(row, "status", "") or "").strip() or "Найден акт",
            notes=str(get_alias(row, "notes", "") or "").strip(),
            lat=parse_float_or_none(get_alias(row, "lat", "")),
            lon=parse_float_or_none(get_alias(row, "lon", "")),
            project_slug=str(get_alias(row, "project_slug", "") or "").strip(),
            source_file=source_file,
        )
        rec.ensure_uid(i)
        if rec.address or rec.object_name or rec.act_url:
            recs.append(rec)
    return recs


def write_records_json(path: str | Path, records: Sequence[HouseRecord]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with p.open("w", encoding="utf-8") as f:
        json.dump({"records": [r.to_json() for r in records]}, f, ensure_ascii=False, indent=2)


def write_records_csv(path: str | Path, records: Sequence[HouseRecord]) -> None:
    p = Path(path)
    p.parent.mkdir(parents=True, exist_ok=True)
    fields = ["uid","address","district","object_name","act_title","act_date","act_url","priority","status","notes","lat","lon","project_slug","source_file"]
    with p.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, delimiter=";")
        w.writeheader()
        for r in records:
            w.writerow({k: getattr(r, k) for k in fields})


def create_house_project(record: HouseRecord, root: str | Path, act_file: str | Path | None = None) -> Dict[str, Path]:
    record.ensure_uid()
    root = Path(root)
    project_dir = root / (record.project_slug or safe_slug(record.address or record.uid))
    paths = {
        "project_dir": project_dir,
        "acts_dir": project_dir / "acts",
        "historical_dir": project_dir / "historical_sources",
        "modern_dir": project_dir / "modern_sources",
        "result_dir": project_dir / "result",
        "annotations_dir": project_dir / "annotations",
        "roboflow_dir": project_dir / "roboflow_export",
        "metadata_dir": project_dir / "metadata",
    }
    for p in paths.values():
        p.mkdir(parents=True, exist_ok=True)
    with (paths["metadata_dir"] / "house.json").open("w", encoding="utf-8") as f:
        json.dump(record.to_json(), f, ensure_ascii=False, indent=2)
    if record.act_url:
        (paths["acts_dir"] / "act_url.txt").write_text(record.act_url + "\n", encoding="utf-8")
        (paths["acts_dir"] / "act_link.url").write_text(f"[InternetShortcut]\nURL={record.act_url}\n", encoding="utf-8")
    if act_file:
        src = Path(act_file)
        if src.exists() and src.is_file():
            shutil.copy2(src, paths["acts_dir"] / src.name)
    (project_dir / "README_HOUSE.md").write_text(make_house_readme(record), encoding="utf-8")
    return paths


def make_house_readme(record: HouseRecord) -> str:
    return f"""# {record.address or record.display_title}

## Объект
{record.object_name or "—"}

## Акт / источник
- Дата / период: {record.act_date or "—"}
- Тема: {record.act_title or "—"}
- Ссылка: {record.act_url or "—"}

## Координаты
- Широта: {record.lat if record.lat is not None else "—"}
- Долгота: {record.lon if record.lon is not None else "—"}

## Статус
{record.status or "Найден акт"}

## Комментарий
{record.notes or "—"}

## Папки
- `acts` — акты и ссылки на документы
- `historical_sources` — исторические фото
- `modern_sources` — современные фото
- `result` — overlay, разметка, итоговые изображения
- `annotations` — JSON/COCO разметка
- `roboflow_export` — экспорт для Roboflow
"""


def search_records(records: Sequence[HouseRecord], query: str = "", status: str = "") -> List[HouseRecord]:
    q = query.strip().lower().replace("ё", "е")
    out: List[HouseRecord] = []
    for r in records:
        if status and r.status != status:
            continue
        text = " ".join([r.address, r.district, r.object_name, r.act_title, r.act_date, r.priority, r.status, r.notes, r.act_url]).lower().replace("ё", "е")
        if not q or q in text:
            out.append(r)
    return out


def open_system_path(path: str | Path) -> None:
    p = Path(path)
    try:
        if sys.platform.startswith("win"):
            os.startfile(str(p))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(p)])
        else:
            subprocess.Popen(["xdg-open", str(p)])
    except Exception:
        pass


class HouseDatabaseFrame(ttk.Frame):
    """Готовая вкладка базы домов для встраивания в Archiview CV."""
    def __init__(self, parent: tk.Widget, project_root: str | Path, on_house_selected: Optional[Callable[[HouseRecord, Dict[str, Path]], None]] = None, on_log: Optional[Callable[[str], None]] = None) -> None:
        super().__init__(parent)
        self.project_root = Path(project_root)
        self.on_house_selected = on_house_selected
        self.on_log = on_log
        self.records: List[HouseRecord] = []
        self.filtered: List[HouseRecord] = []
        self.selected: Optional[HouseRecord] = None
        self.database_path: Optional[Path] = None
        self.search_var = tk.StringVar(value="")
        self.status_var = tk.StringVar(value="")
        self.project_root_var = tk.StringVar(value=str(self.project_root))
        self._build_ui()

    def _build_ui(self) -> None:
        self.columnconfigure(0, weight=1)
        self.rowconfigure(1, weight=1)
        top = ttk.LabelFrame(self, text="База домов / актов ГИКЭ")
        top.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 6))
        top.columnconfigure(1, weight=1)
        ttk.Label(top, text="Корневая папка проектов:").grid(row=0, column=0, sticky="w", padx=8, pady=6)
        ttk.Entry(top, textvariable=self.project_root_var).grid(row=0, column=1, sticky="ew", padx=6, pady=6)
        ttk.Button(top, text="Выбрать…", command=self.choose_project_root).grid(row=0, column=2, padx=8, pady=6)
        btns = ttk.Frame(top)
        btns.grid(row=1, column=0, columnspan=3, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(btns, text="Импорт Excel/CSV", command=self.import_table).pack(side="left", padx=(0, 6))
        ttk.Button(btns, text="Загрузить JSON", command=self.load_json).pack(side="left", padx=(0, 6))
        ttk.Button(btns, text="Сохранить JSON", command=self.save_json).pack(side="left", padx=(0, 6))
        ttk.Button(btns, text="Экспорт CSV", command=self.export_csv).pack(side="left", padx=(0, 6))
        ttk.Button(btns, text="Создать папку и использовать дом", command=self.use_selected_house).pack(side="right")
        filt = ttk.Frame(top)
        filt.grid(row=2, column=0, columnspan=3, sticky="ew", padx=8, pady=(0, 8))
        filt.columnconfigure(1, weight=1)
        ttk.Label(filt, text="Поиск:").grid(row=0, column=0, sticky="w")
        ent = ttk.Entry(filt, textvariable=self.search_var)
        ent.grid(row=0, column=1, sticky="ew", padx=6)
        ent.bind("<KeyRelease>", lambda _e: self.apply_filter())
        ttk.Label(filt, text="Статус:").grid(row=0, column=2, padx=(8, 2))
        box = ttk.Combobox(filt, textvariable=self.status_var, values=[""] + STATUS_OPTIONS, state="readonly", width=24)
        box.grid(row=0, column=3)
        box.bind("<<ComboboxSelected>>", lambda _e: self.apply_filter())

        pane = ttk.Panedwindow(self, orient="vertical")
        pane.grid(row=1, column=0, sticky="nsew", padx=8, pady=6)
        table = ttk.Frame(pane)
        details = ttk.LabelFrame(pane, text="Карточка выбранного дома")
        pane.add(table, weight=3)
        pane.add(details, weight=1)
        table.rowconfigure(0, weight=1)
        table.columnconfigure(0, weight=1)
        cols = ("address", "object", "date", "priority", "status")
        self.tree = ttk.Treeview(table, columns=cols, show="headings", selectmode="browse")
        headers = {"address":"Адрес", "object":"Объект", "date":"Дата / период", "priority":"Приоритет", "status":"Статус"}
        widths = {"address":280, "object":420, "date":120, "priority":110, "status":160}
        for c in cols:
            self.tree.heading(c, text=headers[c])
            self.tree.column(c, width=widths[c], anchor="w")
        y = ttk.Scrollbar(table, orient="vertical", command=self.tree.yview)
        x = ttk.Scrollbar(table, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=y.set, xscrollcommand=x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        y.grid(row=0, column=1, sticky="ns")
        x.grid(row=1, column=0, sticky="ew")
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Double-1>", lambda _e: self.use_selected_house())
        details.columnconfigure(0, weight=1)
        details.rowconfigure(0, weight=1)
        self.details = tk.Text(details, height=8, wrap="word")
        self.details.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)
        ds = ttk.Scrollbar(details, orient="vertical", command=self.details.yview)
        ds.grid(row=0, column=1, sticky="ns", pady=8)
        self.details.configure(yscrollcommand=ds.set)
        btm = ttk.Frame(details)
        btm.grid(row=1, column=0, columnspan=2, sticky="ew", padx=8, pady=(0, 8))
        ttk.Button(btm, text="Скопировать ссылку на акт", command=self.copy_act_url).pack(side="left", padx=(0, 6))
        ttk.Button(btm, text="Открыть папку проекта", command=self.open_selected_project).pack(side="left", padx=(0, 6))
        ttk.Button(btm, text="Создать папку и использовать дом", command=self.use_selected_house).pack(side="right")
        self._set_details("Импортируйте Excel/CSV со списком домов или загрузите сохранённую базу JSON.")

    def log(self, text: str) -> None:
        if self.on_log:
            try:
                self.on_log(text)
            except Exception:
                pass

    def choose_project_root(self) -> None:
        folder = filedialog.askdirectory(title="Выберите корневую папку проектов домов", initialdir=str(self.project_root))
        if folder:
            self.project_root = Path(folder)
            self.project_root_var.set(str(self.project_root))

    def import_table(self) -> None:
        path = filedialog.askopenfilename(title="Импорт списка домов", filetypes=[("Excel / CSV / JSON", "*.xlsx *.xlsm *.csv *.json"), ("Все файлы", "*.*")])
        if not path:
            return
        try:
            self.records = read_house_table(path)
            self.database_path = Path(path)
            self.apply_filter()
            self.log(f"Импортировано домов: {len(self.records)} из {path}\n")
            messagebox.showinfo("Импорт завершён", f"Импортировано записей: {len(self.records)}")
        except Exception as exc:
            messagebox.showerror("Ошибка импорта", str(exc))

    def load_json(self) -> None:
        path = filedialog.askopenfilename(title="Загрузить базу JSON", filetypes=[("JSON", "*.json"), ("Все файлы", "*.*")])
        if not path:
            return
        try:
            self.records = read_house_table(path)
            self.database_path = Path(path)
            self.apply_filter()
            self.log(f"Загружена база: {path}\n")
        except Exception as exc:
            messagebox.showerror("Ошибка загрузки", str(exc))

    def save_json(self) -> None:
        if not self.records:
            messagebox.showwarning("Нет данных", "Сначала импортируйте список домов.")
            return
        default = self.database_path.with_suffix(".json") if self.database_path else Path(self.project_root_var.get()) / "house_database.json"
        path = filedialog.asksaveasfilename(title="Сохранить базу JSON", defaultextension=".json", initialdir=str(default.parent), initialfile=default.name, filetypes=[("JSON", "*.json")])
        if path:
            write_records_json(path, self.records)
            self.database_path = Path(path)
            self.log(f"База сохранена: {path}\n")

    def export_csv(self) -> None:
        if not self.records:
            messagebox.showwarning("Нет данных", "Сначала импортируйте список домов.")
            return
        default = self.database_path.with_suffix(".csv") if self.database_path else Path(self.project_root_var.get()) / "house_database.csv"
        path = filedialog.asksaveasfilename(title="Экспорт CSV", defaultextension=".csv", initialdir=str(default.parent), initialfile=default.name, filetypes=[("CSV", "*.csv")])
        if path:
            write_records_csv(path, self.records)
            self.log(f"CSV сохранён: {path}\n")

    def apply_filter(self) -> None:
        self.filtered = search_records(self.records, self.search_var.get(), self.status_var.get())
        self.tree.delete(*self.tree.get_children())
        for i, r in enumerate(self.filtered):
            self.tree.insert("", "end", iid=str(i), values=(r.address, r.object_name, r.act_date, r.priority, r.status))
        self._set_details(f"Показано записей: {len(self.filtered)} из {len(self.records)}")

    def _on_tree_select(self, _event: object = None) -> None:
        sel = self.tree.selection()
        if not sel:
            self.selected = None
            return
        try:
            self.selected = self.filtered[int(sel[0])]
            self._set_details(self._format_record(self.selected))
        except Exception:
            self.selected = None

    def _set_details(self, text: str) -> None:
        self.details.configure(state="normal")
        self.details.delete("1.0", "end")
        self.details.insert("1.0", text)
        self.details.configure(state="disabled")

    def _format_record(self, r: HouseRecord) -> str:
        return (f"Адрес: {r.address}\nОбъект: {r.object_name}\nТема акта: {r.act_title}\nДата / период: {r.act_date}\nСсылка: {r.act_url}\nПриоритет: {r.priority}\nСтатус: {r.status}\nКоординаты: {r.lat if r.lat is not None else '—'}, {r.lon if r.lon is not None else '—'}\nПапка проекта: {r.project_slug}\n\nКомментарий:\n{r.notes}\n")

    def copy_act_url(self) -> None:
        if not self.selected or not self.selected.act_url:
            messagebox.showwarning("Нет ссылки", "У выбранного дома нет ссылки на акт.")
            return
        self.clipboard_clear()
        self.clipboard_append(self.selected.act_url)
        self.log("Ссылка на акт скопирована в буфер.\n")

    def _selected_project_paths(self) -> Optional[Dict[str, Path]]:
        if not self.selected:
            messagebox.showwarning("Дом не выбран", "Выберите дом в таблице.")
            return None
        self.project_root = Path(self.project_root_var.get() or self.project_root)
        return create_house_project(self.selected, self.project_root)

    def open_selected_project(self) -> None:
        paths = self._selected_project_paths()
        if paths:
            open_system_path(paths["project_dir"])

    def use_selected_house(self) -> None:
        paths = self._selected_project_paths()
        if not paths or not self.selected:
            return
        self.log(f"Создана/выбрана папка проекта дома: {paths['project_dir']}\n")
        if self.on_house_selected:
            self.on_house_selected(self.selected, paths)
        messagebox.showinfo("Дом выбран", "Дом выбран для работы. Переходите к вкладке «Источники».")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("Тест: База домов Archiview CV")
    HouseDatabaseFrame(root, project_root=Path.cwd() / "archiview_projects", on_log=print).pack(fill="both", expand=True)
    root.geometry("1100x760")
    root.mainloop()
